import io
import os
import socket
import datetime
import locale
import zipfile
import asyncio
import re
import uuid
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, UploadFile, File, Request, Form, BackgroundTasks
import glob
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, PlainTextResponse
from fastapi.templating import Jinja2Templates
import pandas as pd
import calendar
import uvicorn
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


from fastapi.staticfiles import StaticFiles


# Setup templates
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

app = FastAPI()

# Mount static files
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")

templates = Jinja2Templates(directory=TEMPLATES_DIR)

# Global Job Store
# Global Job Store
jobs: Dict[str, Any] = {}
MAX_LOG_ENTRIES = 500
PDF_CONCURRENCY = max(1, int(os.environ.get("PDF_CONCURRENCY", "1")))

# Italian month mapping
MONTH_MAP = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
    "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12
}

def parse_italian_date(date_str: str) -> Optional[datetime.date]:
    """
    Parses 'martedì 1 aprile 2025' or '1 aprile 2025' into a date object.
    Ignores weekday if present.
    """
    if not isinstance(date_str, str):
        return pd.to_datetime(date_str).date() if pd.notna(date_str) else None
        
    date_str = date_str.lower().strip()
    
    # Remove weekday if present (assumes simple structure: 'weekday d month y')
    # Use regex to find pattern: day (digits) month (letters) year (4 digits)
    match = re.search(r'(\d+)\s+([a-z]+)\s+(\d{4})', date_str)
    if not match:
        return None
        
    day_s, month_s, year_s = match.groups()
    try:
        day = int(day_s)
        year = int(year_s)
        month = MONTH_MAP.get(month_s)
        if not month:
            return None
        return datetime.date(year, month, day)
    except:
        return None

def get_easter_monday(year):
    """Calculates Easter Monday for a given year."""
    # Anonymous algorithm for Easter date
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    
    easter = datetime.date(year, month, day)
    return easter + datetime.timedelta(days=1)

def get_italian_holidays(year):
    """Returns a set of holiday dates for the year."""
    fixed = [
        (1, 1), (1, 6), (4, 25), (5, 1), (6, 2),
        (8, 15), (11, 1), (12, 8), (12, 25), (12, 26)
    ]
    holidays = {datetime.date(year, m, d) for m, d in fixed}
    holidays.add(get_easter_monday(year))
    return holidays

def preprocess_df(df: pd.DataFrame) -> pd.DataFrame:
    """Cleans and prepares the DataFrame for storage/processing."""
    # Normalize columns
    df.columns = [c.strip() for c in df.columns]
    
    # Required columns check
    required = ["Date", "Name", "Plans", "Hour", "Payroll Number", "Cost Center"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")
    
    # Date processing
    df['DateObj'] = df['Date'].apply(parse_italian_date)
    df = df.dropna(subset=['DateObj']) # Drop invalid dates
    # Convert date objects to timestamp for Parquet compatibility
    df['DateObj'] = pd.to_datetime(df['DateObj'])
    
    # Hour processing
    def parse_hour(h):
        if isinstance(h, (int, float)):
            return float(h)
        if isinstance(h, str):
            try:
                return float(h.replace(',', '.'))
            except:
                return 0.0
        return 0.0
        
    df['HourVal'] = df['Hour'].apply(parse_hour)
    
    return df

def append_job_log(job_id: str, message: str) -> None:
    job = jobs.get(job_id)
    if not job:
        return
    logs = job.setdefault("logs", [])
    logs.append(message)
    if len(logs) > MAX_LOG_ENTRIES:
        del logs[:-MAX_LOG_ENTRIES]

def sanitize_path_component(value: Any, fallback: str) -> str:
    text = str(value or "").strip()
    text = "".join(c for c in text if c.isalnum() or c in (" ", "_", "-")).strip()
    return text or fallback

def build_excel_bytes(user_name: str, user_data: Dict[str, Any]) -> bytes:
    workbook = Workbook()
    first_sheet = True

    header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
    shaded_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    total_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    white_bold_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    for index, month_data in enumerate(user_data.get("blocks", []), start=1):
        if first_sheet:
            ws = workbook.active
            ws.title = (month_data["month_name"] or f"Sheet{index}")[:31]
            first_sheet = False
        else:
            ws = workbook.create_sheet(title=(month_data["month_name"] or f"Sheet{index}")[:31])

        ws["A1"] = "User"
        ws["B1"] = user_name
        ws["D1"] = "Payroll Number"
        ws["E1"] = user_data.get("payroll", "")
        ws["G1"] = "Cost Center"
        ws["H1"] = user_data.get("cost_center", "")

        for cell in ("A1", "D1", "G1"):
            ws[cell].font = bold_font

        header_row = 3
        ws.cell(row=header_row, column=1, value=month_data["month_name"] or f"Month {index}")
        ws.cell(row=header_row, column=1).fill = header_fill
        ws.cell(row=header_row, column=1).font = white_bold_font
        ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left")

        for day in range(1, 32):
            cell = ws.cell(row=header_row, column=day + 1, value=day if day <= month_data["days_in_month"] else "")
            cell.fill = header_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center")

        total_header = ws.cell(row=header_row, column=33, value="Tot")
        total_header.fill = header_fill
        total_header.font = white_bold_font
        total_header.alignment = Alignment(horizontal="center")

        row_idx = header_row + 1
        ws.cell(row=row_idx, column=1, value="Totale")
        ws.cell(row=row_idx, column=1).font = bold_font
        ws.cell(row=row_idx, column=1).fill = total_fill

        for day_idx, cell_val in enumerate(month_data["total_row_cells"], start=2):
            cell = ws.cell(row=row_idx, column=day_idx, value=cell_val)
            cell.alignment = Alignment(horizontal="center")
            if month_data["total_row_day_shaded_status"][day_idx - 2]:
                cell.fill = shaded_fill
            else:
                cell.fill = total_fill

        ws.cell(row=row_idx, column=33, value=month_data["grand_total"])
        ws.cell(row=row_idx, column=33).font = bold_font
        ws.cell(row=row_idx, column=33).fill = total_fill
        ws.cell(row=row_idx, column=33).alignment = Alignment(horizontal="right")

        for row in month_data["plans"]:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=row["plan_name"])
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

            for day_offset, day_val in enumerate(row["days"], start=2):
                value = "" if not day_val["date_valid"] else day_val["value"]
                cell = ws.cell(row=row_idx, column=day_offset, value=value)
                cell.alignment = Alignment(horizontal="center")
                if day_val["shaded"]:
                    cell.fill = shaded_fill

            ws.cell(row=row_idx, column=33, value=row["total"])
            ws.cell(row=row_idx, column=33).alignment = Alignment(horizontal="right")

        ws.freeze_panes = "B4"
        ws.column_dimensions["A"].width = 36
        for col_idx in range(2, 33):
            ws.column_dimensions[get_column_letter(col_idx)].width = 5
        ws.column_dimensions["AG"].width = 10

    if first_sheet:
        ws = workbook.active
        ws.title = "Timesheet"
        ws["A1"] = "Nessun dato disponibile"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def build_users_dict(df: pd.DataFrame) -> Dict[str, Any]:
    """Organizes the CLEAN dataframe into the report structure."""
    users_data = {}
    for name, user_df in df.groupby("Name"):
        # Payroll and Cost Center (take first valid)
        payroll = user_df["Payroll Number"].iloc[0] if not user_df["Payroll Number"].empty else ""
        try:
            if pd.notna(payroll):
                # Handle string numbers with commas
                payroll_str = str(payroll).replace(',', '.')
                payroll = str(int(float(payroll_str)))
        except:
            payroll = str(payroll)

        cost_center = user_df["Cost Center"].iloc[0] if not user_df["Cost Center"].empty else ""
        if pd.isna(cost_center):
            cost_center = "Unknown"
        
        # Sort by date
        user_df = user_df.sort_values('DateObj')
        
        # Identify months needed
        unique_periods = user_df['DateObj'].apply(lambda d: (d.year, d.month)).drop_duplicates()
        
        month_blocks = []
        
        for (year, month) in unique_periods:
            # Filter for this month
            month_df = user_df[
                (user_df['DateObj'].dt.year == year) & 
                (user_df['DateObj'].dt.month == month)
            ].copy()
            
            # Build matrix
            month_df['Day'] = month_df['DateObj'].dt.day
            
            pivot = month_df.pivot_table(
                index="Plans", 
                columns="Day", 
                values="HourVal", 
                aggfunc="sum",
                fill_value=0
            )
            
            # Ensure columns 1..31 exist
            for d in range(1, 32):
                if d not in pivot.columns:
                    pivot[d] = 0
            
            # Reorder columns 1..31
            pivot = pivot[sorted(pivot.columns)]
            
            # Rows processing
            rows = []
            
            # Calculate Totals
            pivot['Total'] = pivot.sum(axis=1)
            total_row_vals = pivot.sum(axis=0)
            
            # Identify month name
            # month_name_it = [k for k,v in MONTH_MAP.items() if v == month][0]
            # MAPPING FOR ENGLISH UPPERCASE
            MONTH_MAP_EN = {
                1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL", 5: "MAY", 6: "JUNE",
                7: "JULY", 8: "AUGUST", 9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER"
            }
            month_name_it = MONTH_MAP_EN.get(month, "")
            
            # Holidays for this year
            holidays = get_italian_holidays(year)
            
            day_meta = {}
            for d in range(1, 32):
                date_obj = None
                try:
                    date_obj = datetime.date(year, month, d)
                except ValueError:
                    pass
                
                is_holiday = False
                is_weekend = False
                if date_obj:
                    if date_obj.weekday() >= 5: # 5=Sat, 6=Sun
                        is_weekend = True
                    if date_obj in holidays:
                        is_holiday = True
                
                day_meta[d] = {
                    "day": d,
                    "shaded": is_weekend or is_holiday,
                    "date_valid": date_obj is not None
                }
            
            # Format numbers: 2 decimals, comma separator
            def fmt(n, is_zero_target_empty=True):
                if n == 0 and is_zero_target_empty:
                    return ""
                s = "{:.2f}".format(n).replace('.', ',')
                return s
            
            # Total Row
            total_row_cells = []
            total_row_day_values = []
            total_row_day_shaded_status = []
            for d in range(1, 32):
                val = total_row_vals[d]
                meta = day_meta[d]
                total_row_day_values.append(val)
                total_row_day_shaded_status.append(meta["shaded"])
            
            grand_total = total_row_vals['Total']

            # Get days in month for header hiding logic
            days_in_month = calendar.monthrange(year, month)[1]

            # Plan Rows
            plans_list = []
            for plan_name, row_series in pivot.iterrows():
                if plan_name == "Total": continue
                
                row_cells = []
                for d in range(1, 32):
                    val = row_series[d]
                    meta = day_meta[d]
                    row_cells.append({
                        "value": fmt(val, is_zero_target_empty=True),
                        "shaded": meta["shaded"],
                        "date_valid": meta["date_valid"]
                    })
                
                plans_list.append({
                    "plan_name": plan_name,
                    "is_total": False,
                    "is_indent": str(plan_name).startswith("_"),
                    "days": row_cells,
                    "total": fmt(row_series['Total'], is_zero_target_empty=True)
                })

            month_blocks.append({
                "month_name": month_name_it,
                "days_in_month": days_in_month,
                "plans": plans_list,
                "total_row_cells": [fmt(x, is_zero_target_empty=True) for x in total_row_day_values],
                "total_row_day_shaded_status": total_row_day_shaded_status, # Add shaded status for total row
                "grand_total": fmt(grand_total, is_zero_target_empty=True)
            })
            
        users_data[name] = {
            "payroll": payroll,
            "cost_center": cost_center,
            "blocks": month_blocks
        }
        
    return users_data

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/robots.txt", response_class=PlainTextResponse)
async def robots():
    return "User-agent: *\\nDisallow: /"

async def process_generation_task(job_id: str, filtered_data: dict, output_pdf: bool, output_excel: bool):
    jobs[job_id]["status"] = "processing"
    jobs[job_id]["progress"] = 0
    jobs[job_id]["total"] = len(filtered_data)
    
    zip_path = os.path.join(BASE_DIR, f"output_{job_id}.zip")

    try:
        sem = asyncio.Semaphore(PDF_CONCURRENCY)
        completed_pdfs = 0
        
        append_job_log(job_id, f"[SYSTEM] Initializing browser engine...")
        
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            async with async_playwright() as p:
                browser = await p.chromium.launch(args=["--disable-dev-shm-usage", "--no-sandbox"])
            
                async def generate_single_user_files(user_name, user_data):
                    page = None
                    async with sem:
                        try:
                            append_job_log(job_id, f"[TASK] Generating files for {user_name}...")
                            archive_entries = []

                            if output_pdf:
                                html_content = templates.get_template("timesheet_report.html").render(
                                    user_name=user_name,
                                    payroll_number=user_data["payroll"],
                                    cost_center=user_data["cost_center"],
                                    blocks=user_data["blocks"],
                                    is_shaded_col=lambda d: False
                                )

                                footer_html = """
        <div style="font-size: 8px; font-family: sans-serif; width: 100%; display: flex; justify-content: space-between; padding: 0 20px; color: #555;">
            <span>Year: <span class="date"></span></span>
            <span>Page <span class="pageNumber"></span> / <span class="totalPages"></span></span>
            <span>Signature: __________________________</span>
        </div>
        """

                                page = await browser.new_page()
                                await page.set_content(html_content, wait_until="load")

                                pdf_bytes = await page.pdf(
                                    format="A4",
                                    print_background=True,
                                    landscape=False,
                                    display_header_footer=True,
                                    header_template="<div></div>",
                                    footer_template=footer_html,
                                    margin={"top": "1cm", "right": "0.5cm", "bottom": "1.5cm", "left": "0.5cm"}
                                )
                                archive_entries.append(("pdf", pdf_bytes))
                                await page.close()
                                page = None

                            if output_excel:
                                excel_bytes = build_excel_bytes(user_name, user_data)
                                archive_entries.append(("xlsx", excel_bytes))

                            cc_folder = sanitize_path_component(user_data.get("cost_center"), "Unknown")
                            safe_name = sanitize_path_component(user_name, "utente")
                            base_archive_name = f"{cc_folder}/{safe_name}"

                            append_job_log(job_id, f"[SUCCESS] {user_name} completed.")
                            return [(f"{base_archive_name}.{ext}", file_bytes) for ext, file_bytes in archive_entries]
                        except Exception as e:
                            append_job_log(job_id, f"[ERROR] Failed {user_name}: {str(e)}")
                            return None
                        finally:
                            if page is not None:
                                await page.close()

                tasks = [
                    asyncio.create_task(generate_single_user_files(name, data))
                    for name, data in filtered_data.items()
                ]
            
                for task in asyncio.as_completed(tasks):
                    res = await task
                    if not res:
                        continue
                    for archive_name, file_bytes in res:
                        zf.writestr(archive_name, file_bytes)
                        del file_bytes
                    completed_pdfs += 1
                    jobs[job_id]["progress"] = completed_pdfs
            
        if completed_pdfs == 0:
            raise RuntimeError("Nessun file generato correttamente.")

        jobs[job_id]["filename"] = zip_path
        jobs[job_id]["status"] = "completed"
        append_job_log(job_id, f"[SYSTEM] Job Dispatched. Ready for download.")
        
    except Exception as e:
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["error"] = str(e)
        append_job_log(job_id, f"[CRITICAL] Job Failed: {str(e)}")
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
        except OSError:
            pass

@app.post("/generate")
async def generate_pdf(
    background_tasks: BackgroundTasks,
    request: Request,
    cost_centers: List[str] = Form(...),
    file_id: str = Form(...),
    output_pdf: Optional[str] = Form(None),
    output_excel: Optional[str] = Form(None)
):
    temp_file = f"temp_{file_id}.parquet"
    if not os.path.exists(temp_file):
        return JSONResponse({"error": "Sessione scaduta o file non trovato"}, status_code=400)
        
    try:
        try:
            df = pd.read_parquet(temp_file)
        finally:
            try:
                os.remove(temp_file)
            except OSError:
                pass
        
        try:
            data = build_users_dict(df)
        except Exception as e:
             return JSONResponse({"error": f"Errore elaborazione: {str(e)}"}, status_code=400)
        
        # Filter users based on selected Cost Centers
        filtered_data = {
            name: u_data 
            for name, u_data in data.items() 
            if u_data.get("cost_center") in cost_centers or (not u_data.get("cost_center") and "Unknown" in cost_centers)
        }
        
        if not filtered_data:
             return JSONResponse({"error": "Nessun utente trovato"}, status_code=400)

        wants_pdf = output_pdf is not None
        wants_excel = output_excel is not None
        if not wants_pdf and not wants_excel:
            return JSONResponse({"error": "Seleziona almeno un formato di output"}, status_code=400)

        # Create Job
        job_id = str(uuid.uuid4())
        jobs[job_id] = {
            "status": "pending",
            "progress": 0,
            "total": len(filtered_data),
            "logs": [],
            "filename": None
        }
        
        # Start Background Task
        background_tasks.add_task(process_generation_task, job_id, filtered_data, wants_pdf, wants_excel)
        
        return JSONResponse({"job_id": job_id})
    
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/job/{job_id}")
async def get_job_status(job_id: str):
    if job_id not in jobs:
        return JSONResponse({"error": "Job not found"}, status_code=404)
    return JSONResponse(jobs[job_id])

@app.get("/download/{job_id}")
async def download_job(job_id: str, background_tasks: BackgroundTasks):
    if job_id not in jobs or jobs[job_id]["status"] != "completed":
         return JSONResponse({"error": "File not ready"}, status_code=404)
         
    path = jobs[job_id]["filename"]
    if not path or not os.path.exists(path):
        return JSONResponse({"error": "File not found"}, status_code=404)
    
    def cleanup():
        try:
            os.remove(path)
            del jobs[job_id]
        except:
            pass
            
    background_tasks.add_task(cleanup)
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"timesheet_pdfs_{timestamp}.zip"
    
    return StreamingResponse(
        open(path, "rb"), 
        media_type="application/zip", 
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.post("/analyze", response_class=HTMLResponse)
async def analyze_excel(request: Request, file: UploadFile = File(...)):
    content = await file.read()
    
    try:
        # Read Excel ONCE
        df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
        
        # Preprocess NOW to ensure clean types for Parquet
        df = preprocess_df(df)
        
        # Save CLEAN data to Unique Parquet
        file_id = str(uuid.uuid4())
        temp_file = f"temp_{file_id}.parquet"
        df.to_parquet(temp_file, engine='pyarrow')
        
        # Cost Center counting
        if "Cost Center" in df.columns:
            # Group by Name to count users, not rows (assuming one cost center per user)
            # Use 'first' to get a representative cost center for the user
            users_cc = df.groupby("Name")["Cost Center"].first()
            # Handle potential None/NaN values in cost center
            users_cc = users_cc.fillna("Unknown")
            cost_centers = users_cc.value_counts().to_dict()
        else:
            cost_centers = {"Unknown": df["Name"].nunique()}

        return templates.TemplateResponse("select_cost_centers.html", {
            "request": request,
            "cost_centers": cost_centers,
            "file_id": file_id
        })
        
    except Exception as e:
         return HTMLResponse(f"<h3>Errore nella lettura del file:</h3><p>{str(e)}</p>", status_code=400)

if __name__ == "__main__":
    # Startup cleanup
    for f in glob.glob("temp*.parquet"):
        try:
            os.remove(f)
        except:
            pass
    for f in glob.glob("output_*.zip"):
        try:
            os.remove(f)
        except:
            pass
            
    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            # Doesn't need to be reachable
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"

    local_ip = get_local_ip()
    print(f"\\n--- APP PRONTA ---")
    print(f"Locale:  http://localhost:8108")
    print(f"Rete:    http://{local_ip}:8108")
    print(f"------------------\\n")

    uvicorn.run(app, host="0.0.0.0", port=8108)
